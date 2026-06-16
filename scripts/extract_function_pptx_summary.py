#!/usr/bin/env python
from __future__ import annotations

import argparse
import csv
import re
import struct
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from inspect_ole_cfb import CfbFile, clean_text


TEXT_NS = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
LABEL_ATOM_PATTERN = r"[A-Za-z0-9_()+%.\-\u4e00-\u9fffμµ]+"
CONTROL_LABELS = {
    "no antibody no emig",
    "wtih antibody no emig",
    "with antibody no emig",
    "1 μg plasmid",
    "1μg plasmid",
    "1 ug plasmid",
    "1ug plasmid",
}
LABELS = [
    "no antibody no emig",
    "wtih antibody no emig",
    "with antibody no emig",
    "1 μg plasmid",
    "1μg plasmid",
    "1 ug plasmid",
    "1ug plasmid",
    "混合后eMig",
    "预组装beads",
    "eMig直接加Apo",
    "20um一级",
    "SD-3+0.8",
    "SD-1.2",
    "90%-400RPM",
    "70%-200RPM",
    "60%-200RPM",
    "50%-100RPM",
    "NC-10ug",
    "NC-6ug",
    "NC-2ug",
    "MK-2.5",
    "MK-1.2",
    "eC-eM",
    "SD-8",
    "SD-5",
    "SD-3",
    "MK-5",
    "TRE-effect-eMig",
    "PBS-速冻",
    "PBS-慢冻",
    "TRE-速冻",
    "TRE-慢冻",
    "PBS-fake",
    "未处理组",
    "(NC)90%-400RPM",
    "(NC)90%-600RPM",
    "(NC)80%-400RPM",
    "(NC)80%-600RPM",
    "F8+9",
    "F8-9",
    "F10",
    "F9",
    "F8",
    "F7",
    "F6",
    "F5",
    "F4",
    "F3",
    "F2",
    "F1",
    "NC",
]


def slide_no(name: str) -> int:
    match = re.search(r"slide(\d+)\.xml$", name)
    return int(match.group(1)) if match else 0


def slide_text(zf: zipfile.ZipFile, slide_xml: str) -> str:
    root = ET.fromstring(zf.read(slide_xml))
    return " | ".join(node.text for node in root.findall(".//a:t", TEXT_NS) if node.text)


def extract_experiment_name(raw_slide_title: str) -> str:
    normalized = re.sub(r"\s*\|\s*", "-", raw_slide_title)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    match = re.search(r"(DWD-.+?)-王硕(?:-\d{8})?", normalized)
    if match:
        return match.group(1).strip("- ")
    match = re.search(r"(DWD-.+?)(?:-敲除率|-细胞占比|-保密信息|$)", normalized)
    if match:
        return match.group(1).strip("- ")
    return raw_slide_title


def slide_ole_targets(zf: zipfile.ZipFile, no: int) -> list[str]:
    rel_name = f"ppt/slides/_rels/slide{no}.xml.rels"
    if rel_name not in zf.namelist():
        return []
    xml = zf.read(rel_name).decode("utf-8", errors="ignore")
    return ["ppt/embeddings/" + target for target in re.findall(r'Target="\.\./embeddings/([^"]+\.bin)"', xml)]


def stream_map(blob: bytes) -> dict[str, bytes]:
    cfb = CfbFile(blob)
    out = {}
    for entry in cfb.entries:
        if entry.object_type == 2:
            out[entry.name] = cfb.read_stream(entry)
    return out


def classify_endpoint(streams: dict[str, bytes], visible_slide_text: str = "") -> tuple[str | None, str, str]:
    contents = clean_text(streams.get("CONTENTS", b""))
    preview_text = "\n".join(clean_text(data) for name, data in streams.items() if "OlePres" in name)
    has_function_axis = "敲除" in contents or "knockout" in contents.lower() or "function" in contents.lower()
    has_function_preview = "敲除" in preview_text or "knockout" in preview_text.lower() or "function" in preview_text.lower()
    has_cell_share = "细胞占比" in contents or "viability" in contents.lower() or "population" in contents.lower()
    has_cell_share_preview = "细胞占比" in preview_text or "viability" in preview_text.lower() or "population" in preview_text.lower()
    if has_cell_share:
        return "cell_share_pct", "细胞占比%", "y-axis contains 细胞占比%"
    if has_cell_share_preview:
        return "cell_share_pct", "细胞占比%", "visible chart contains 细胞占比%"
    if has_function_axis or has_function_preview:
        return "knockout_rate_pct", "敲除率（%）", "y-axis/visible chart contains 敲除率"
    if "敲除" in visible_slide_text or "knockout" in visible_slide_text.lower() or "功能" in visible_slide_text:
        if "Prism" in clean_text(streams.get("\x01CompObj", b"")):
            return "knockout_rate_pct", "敲除率（%）", "visible slide title contains 敲除率; OLE is Prism and does not contain 细胞占比"
    return None, "", "excluded: no supported y-axis evidence"


def label_needles(label: str) -> list[bytes]:
    needles = []
    encodings = ("latin1", "utf-8") if label.isascii() else ("utf-8", "gbk", "gb18030")
    for enc in encodings:
        try:
            raw = label.encode(enc, errors="strict")
        except Exception:
            continue
        if raw and raw not in needles:
            needles.append(raw)
    return needles


def find_positions(blob: bytes, needle: bytes) -> list[int]:
    positions = []
    start = 0
    while True:
        idx = blob.find(needle, start)
        if idx < 0:
            return positions
        positions.append(idx)
        start = idx + 1


def looks_embedded_in_longer_label(blob: bytes, idx: int, raw: bytes) -> bool:
    if not re.fullmatch(rb"[A-Za-z]+\d*", raw):
        return False
    nxt = idx + len(raw)
    if nxt >= len(blob):
        return False
    return blob[nxt : nxt + 1] in b"+-0123456789"


def candidate_floats(blob: bytes, idx: int) -> list[tuple[int, float]]:
    vals = []
    for off in range(idx, min(len(blob) - 4, idx + 180)):
        val = struct.unpack("<f", blob[off : off + 4])[0]
        if 0.5 <= val <= 100:
            rel = off - idx
            if not vals or abs(vals[-1][1] - val) > 1e-5 or rel - vals[-1][0] > 8:
                vals.append((rel, val))
    return vals


def label_before_value(blob: bytes, off: int) -> str:
    raw = blob[max(0, off - 96) : off]
    text = raw.decode("gb18030", errors="ignore")
    text = "".join(ch if ch.isprintable() else " " for ch in text)
    tokens = re.findall(LABEL_ATOM_PATTERN, text)
    junk = {
        "MT",
        "ITC",
        "New",
        "BT",
        "BISansCond",
        "Light",
        "Gisha",
        "Sans",
        "Arial",
        "Qa",
        "en",
        "HY",
        "HGS",
        "M-PRO",
        "oPOP",
        "Typewriter",
        "CharterITCReg",
        "RDC.tmp",
        "ZDingbats",
    }
    tokens = [
        tok
        for tok in tokens
        if tok not in junk and (len(tok) >= 2 or re.fullmatch(r"\d+(?:\.\d+)?", tok))
    ]
    if not tokens:
        return ""

    lower_tokens = [tok.lower() for tok in tokens]
    for size in range(min(5, len(tokens)), 1, -1):
        candidate = " ".join(tokens[-size:])
        if is_assay_control(candidate):
            return candidate.strip("- ")

    if lower_tokens[-1] in {"effect"} and len(tokens) >= 2:
        return " ".join(tokens[-2:]).strip("- ")
    if lower_tokens[-1] == "emig" and len(tokens) >= 4 and lower_tokens[-4:-1] in (
        ["no", "antibody", "no"],
        ["with", "antibody", "no"],
        ["wtih", "antibody", "no"],
    ):
        return " ".join(tokens[-4:]).strip("- ")
    if lower_tokens[-1] == "plasmid" and len(tokens) >= 2:
        if re.search(r"(?:\d+\s*)?[μµu]g$", tokens[-2], re.IGNORECASE):
            return " ".join(tokens[-2:]).strip("- ")
        if len(tokens) >= 3 and re.fullmatch(r"\d+", tokens[-3]) and re.fullmatch(r"[μµu]g", tokens[-2], re.IGNORECASE):
            return " ".join(tokens[-3:]).strip("- ")

    return tokens[-1].strip("-")


def plausible_group_label(label: str) -> bool:
    if not label:
        return False
    if len(label) > 48:
        return False
    if not re.fullmatch(rf"{LABEL_ATOM_PATTERN}(?: {LABEL_ATOM_PATTERN})*", label):
        return False
    if not re.search(r"[A-Za-z0-9]", label) and label not in LABELS:
        return False
    junk_fragments = (
        "Arial",
        "Sans",
        "Serif",
        "Wing",
        "Gothic",
        "Times",
        "Calibri",
        "Font",
        "Helvetica",
        "Mono",
        "Book",
        "Italic",
        "Roman",
        "Symbol",
        "Heiti",
        "Hira",
        "222",
        "乮",
        "乯",
        "牋",
        "栽",
        "梃",
        "鱊",
        "览",
        "惪",
        "RDC.tmp",
    )
    if any(fragment in label for fragment in junk_fragments):
        return False
    if label.lower() in {"emig", "plasmid", "ww", "sw", "ic", "en", "qa", "stain", "ff"}:
        return False
    if re.fullmatch(r"\d+[A-Za-z]", label) and not re.fullmatch(r"\d+(?:\.\d+)?h", label, re.IGNORECASE):
        return False
    if re.fullmatch(r"[A-Za-z]+-[\u4e00-\u9fff]{4,}", label):
        return False
    return True


def plausible_single_label(label: str) -> bool:
    return bool(re.search(r"(?:\d+\s*ug|\d+%-\d+RPM|NC-\d+ug)", label, re.IGNORECASE))


def discover_label_values(contents: bytes) -> dict[str, tuple[str, list[float], int]]:
    candidates: list[tuple[int, str, str, list[float]]] = []
    for off in range(115000, len(contents) - 104):
        v1 = struct.unpack("<f", contents[off : off + 4])[0]
        if not 0.5 <= v1 <= 100:
            continue
        label = label_before_value(contents, off)
        if not plausible_group_label(label):
            continue
        v2_48 = struct.unpack("<f", contents[off + 48 : off + 52])[0]
        v3_96 = struct.unpack("<f", contents[off + 96 : off + 100])[0]
        if 0.5 <= v2_48 <= 100 and 0.5 <= v3_96 <= 100:
            candidates.append((off, label, "3-level gradient", [v1, v2_48, v3_96]))
            continue
        v2_12 = struct.unpack("<f", contents[off + 12 : off + 16])[0]
        if 0.5 <= v2_12 <= 100:
            candidates.append((off, label, "two-replicate comparison", [v1, v2_12]))
            continue
        if plausible_single_label(label):
            candidates.append((off, label, "single-value comparison", [v1]))

    if not candidates:
        return {}

    candidates.sort(key=lambda item: item[0])
    clusters: list[list[tuple[int, str, str, list[float]]]] = []
    current: list[tuple[int, str, str, list[float]]] = []
    for candidate in candidates:
        if current and candidate[0] - current[-1][0] > 900:
            clusters.append(current)
            current = []
        current.append(candidate)
    if current:
        clusters.append(current)

    def unique_count(cluster: list[tuple[int, str, str, list[float]]]) -> int:
        return len({item[1] for item in cluster})

    data_clusters = [cluster for cluster in clusters if unique_count(cluster) >= 2]
    selected = data_clusters[0] if data_clusters else clusters[0]

    discovered: dict[str, tuple[str, list[float], int]] = {}
    for off, label, layout, values in selected:
        discovered.setdefault(label, (layout, values, off))
    return discovered


def extract_label_values(contents: bytes, label: str) -> tuple[str, list[float], int] | None:
    best: tuple[str, list[float], int] | None = None
    for raw in label_needles(label):
        for idx in find_positions(contents, raw):
            if looks_embedded_in_longer_label(contents, idx, raw):
                continue
            vals = candidate_floats(contents, idx)
            if not vals:
                continue
            rels = [rel for rel, _ in vals]
            nums = [val for _, val in vals]
            if len(vals) >= 3:
                for pos in range(0, len(vals) - 2):
                    if abs((rels[pos + 1] - rels[pos]) - 48) <= 4 and abs((rels[pos + 2] - rels[pos + 1]) - 48) <= 4:
                        return "3-level gradient", nums[pos : pos + 3], idx
            if len(vals) >= 2:
                for pos in range(0, len(vals) - 1):
                    if abs((rels[pos + 1] - rels[pos]) - 12) <= 3:
                        return "two-replicate comparison", nums[pos : pos + 2], idx
            if vals and 12 <= rels[0] <= 45:
                return "single-value comparison", nums[:1], idx
            if best is None:
                best = ("ambiguous", nums[: min(3, len(nums))], idx)
    return best


def extract_all_label_values(contents: bytes) -> dict[str, tuple[str, list[float], int]]:
    rows = discover_label_values(contents)
    for label in LABELS:
        if label in rows:
            continue
        extracted = extract_label_values(contents, label)
        if extracted:
            rows[label] = extracted
    return rows


def normalized_control_label(label: str) -> str:
    normalized = label.strip().lower()
    normalized = normalized.replace("μ", "u").replace("µ", "u")
    normalized = re.sub(r"\s+", " ", normalized)
    normalized = normalized.replace("1 ug", "1ug")
    return normalized


NORMALIZED_CONTROL_LABELS = {normalized_control_label(label) for label in CONTROL_LABELS}


def is_assay_control(label: str) -> bool:
    return normalized_control_label(label) in NORMALIZED_CONTROL_LABELS


def empty_endpoint_data() -> dict[str, object]:
    return {
        "assay_layout": "",
        "values": [],
        "summary": "",
        "source_ole": "",
        "selection_evidence": "",
        "byte_offset": "",
    }


def summarize_values(layout: str, values: list[float]) -> float | str:
    if not values:
        return ""
    if layout == "3-level gradient" and len(values) >= 3:
        return round(sum(values[:3]) / 3.0, 6)
    if len(values) >= 2:
        return round(sum(values[:2]) / 2.0, 6)
    return round(values[0], 6)


def set_endpoint_columns(row: dict[str, object], prefix: str, data: dict[str, object]) -> None:
    values = data.get("values") or []
    if not isinstance(values, list):
        values = []
    row[f"{prefix}_assay_layout"] = data.get("assay_layout", "")
    for idx in range(3):
        row[f"{prefix}_{idx + 1}_pct"] = round(values[idx], 6) if idx < len(values) else ""
    row[f"{prefix}_summary_pct"] = data.get("summary", "")
    row[f"{prefix}_source_ole"] = data.get("source_ole", "")
    row[f"{prefix}_selection_evidence"] = data.get("selection_evidence", "")
    row[f"{prefix}_byte_offset_in_contents"] = data.get("byte_offset", "")


def write_xlsx(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "extracted_data"
    worksheet.append(fieldnames)

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    control_fill = PatternFill("solid", fgColor="FFF2CC")
    for cell in worksheet[1]:
        cell.fill = header_fill

    for row in rows:
        worksheet.append([row.get(field, "") for field in fieldnames])
        if row.get("is_assay_control") == "TRUE":
            for cell in worksheet[worksheet.max_row]:
                cell.fill = control_fill

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    for col_idx, field in enumerate(fieldnames, start=1):
        max_len = len(field)
        for row_idx in range(2, min(worksheet.max_row, 200) + 1):
            value = worksheet.cell(row_idx, col_idx).value
            if value is not None:
                max_len = max(max_len, len(str(value)))
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 42)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--ppt", required=True)
    parser.add_argument("--out", required=True)
    args = parser.parse_args()

    ppt = Path(args.ppt)
    header = ppt.read_bytes()[:8]
    if not header.startswith(b"PK") or not zipfile.is_zipfile(ppt):
        raise SystemExit("Input is not a readable/decrypted PPTX package; provide a decrypted copy.")

    rows = []
    excluded = []
    with zipfile.ZipFile(ppt) as zf:
        slides = sorted(
            [name for name in zf.namelist() if re.match(r"ppt/slides/slide\d+\.xml$", name)],
            key=slide_no,
        )
        for slide in slides:
            no = slide_no(slide)
            title = slide_text(zf, slide)
            experiment_name = extract_experiment_name(title)
            endpoint_rows: dict[str, dict[str, dict[str, object]]] = {
                "knockout_rate_pct": {},
                "cell_share_pct": {},
            }
            for ole in slide_ole_targets(zf, no):
                streams = stream_map(zf.read(ole))
                endpoint, y_axis_title, evidence = classify_endpoint(streams, title)
                if endpoint is None:
                    excluded.append({"slide": no, "experiment": experiment_name, "source_ole": ole, "reason": evidence})
                    continue
                contents = streams.get("CONTENTS", b"")
                for label, extracted in extract_all_label_values(contents).items():
                    layout, values, offset = extracted
                    endpoint_rows[endpoint][label] = {
                        "assay_layout": layout,
                        "values": values,
                        "summary": summarize_values(layout, values),
                        "source_ole": ole,
                        "selection_evidence": evidence,
                        "byte_offset": offset,
                    }

            all_groups = []
            seen_groups = set()
            for endpoint in ("knockout_rate_pct", "cell_share_pct"):
                for group in endpoint_rows[endpoint]:
                    if group not in seen_groups:
                        all_groups.append(group)
                        seen_groups.add(group)
            knockout_groups = set(endpoint_rows["knockout_rate_pct"])
            cell_share_groups = set(endpoint_rows["cell_share_pct"])
            for group_idx, group in enumerate(all_groups):
                warnings = []
                if group not in knockout_groups:
                    warnings.append("missing knockout_rate_pct")
                if group not in cell_share_groups:
                    warnings.append("missing cell_share_pct")
                row = {
                    "slide": no,
                    "experiment": experiment_name,
                    "group": group,
                    "group_order": group_idx + 1,
                    "is_assay_control": "TRUE" if is_assay_control(group) else "FALSE",
                    "source": "Prism/OLE original data",
                    "data_check_warning": "; ".join(warnings),
                }
                set_endpoint_columns(row, "knockout_rate", endpoint_rows["knockout_rate_pct"].get(group, empty_endpoint_data()))
                set_endpoint_columns(row, "cell_share", endpoint_rows["cell_share_pct"].get(group, empty_endpoint_data()))
                rows.append(row)

    out = Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "slide",
        "experiment",
        "group",
        "group_order",
        "is_assay_control",
        "source",
        "knockout_rate_assay_layout",
        "knockout_rate_1_pct",
        "knockout_rate_2_pct",
        "knockout_rate_3_pct",
        "knockout_rate_summary_pct",
        "knockout_rate_source_ole",
        "knockout_rate_selection_evidence",
        "knockout_rate_byte_offset_in_contents",
        "cell_share_assay_layout",
        "cell_share_1_pct",
        "cell_share_2_pct",
        "cell_share_3_pct",
        "cell_share_summary_pct",
        "cell_share_source_ole",
        "cell_share_selection_evidence",
        "cell_share_byte_offset_in_contents",
        "data_check_warning",
    ]
    rows.sort(
        key=lambda row: (
            row.get("is_assay_control") == "TRUE",
            int(row.get("slide") or 0),
            int(row.get("group_order") or 0),
        )
    )

    csv_out = out if out.suffix.lower() != ".xlsx" else out.with_suffix(".csv")
    xlsx_out = out if out.suffix.lower() == ".xlsx" else out.with_suffix(".xlsx")

    with csv_out.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    write_xlsx(xlsx_out, rows, fieldnames)

    excluded_out = csv_out.with_name(csv_out.stem + "-excluded_ole.csv")
    with excluded_out.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=["slide", "experiment", "source_ole", "reason"])
        writer.writeheader()
        writer.writerows(excluded)
    print(f"Wrote {len(rows)} extracted rows to {csv_out}")
    print(f"Wrote {len(rows)} extracted rows to {xlsx_out}")
    print(f"Wrote {len(excluded)} excluded OLE rows to {excluded_out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
