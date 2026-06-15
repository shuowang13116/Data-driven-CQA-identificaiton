from __future__ import annotations

import argparse
import math
import re
from pathlib import Path
from typing import Iterable

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


PARTICLE_DATA_EXTENSIONS = {".xls", ".xlsx", ".csv", ".tsv", ".txt"}


PDF_FIELD_PATTERNS: dict[str, list[str]] = {
    "mean_diameter_nm_pdf": [
        r"平均粒径[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?)",
        r"Mean\s+Diameter[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?)",
    ],
    "mode_diameter_nm_pdf": [
        r"集中粒径[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?)",
        r"Mode\s+Diameter[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?)",
    ],
    "median_diameter_d50_nm_pdf": [
        r"中位粒径\s*\([^)]*D50[^)]*\)[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?)",
        r"\bD50\b[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?)",
    ],
    "max_diameter_nm_pdf": [
        r"最大粒径[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?)",
        r"Max(?:imum)?\s+Diameter[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?)",
    ],
    "min_diameter_nm_pdf": [
        r"最小粒径[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?)",
        r"Min(?:imum)?\s+Diameter[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?)",
    ],
    "std_diameter_nm_pdf": [
        r"粒径标准差[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?)",
        r"Std(?:\.|andard)?\s+Diameter[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?)",
    ],
    "d90_nm_pdf": [
        r"(?<![A-Za-z0-9])D90(?!\s*[:：]\s*D10)[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?)",
    ],
    "d10_nm_pdf": [
        r"(?<![A-Za-z0-9])D10[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?)",
    ],
    "d90_d10_ratio_pdf": [
        r"D90\s*[:：/]\s*D10[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?)",
    ],
    "span_pdf": [
        r"跨距\s*\([^)]*Span[^)]*\)[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?)",
        r"\bSpan\b[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?)",
    ],
    "dispersion_coefficient_pdf": [
        r"Dispersion\s+Coefficient[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?)",
    ],
    "particle_count_pdf": [
        r"颗粒数[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?)",
        r"Particle\s+Count[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?)",
    ],
    "test_concentration_particles_ml_pdf": [
        r"测试浓度[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?(?:[Ee][+-]?[0-9]+)?)",
        r"Test\s+Concentration[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?(?:[Ee][+-]?[0-9]+)?)",
    ],
    "original_concentration_particles_ml_pdf": [
        r"原始浓度[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?(?:[Ee][+-]?[0-9]+)?)",
        r"Original\s+Concentration[^0-9]{0,40}([0-9]+(?:\.[0-9]+)?(?:[Ee][+-]?[0-9]+)?)",
    ],
}


def clean_pdf_text(text: str) -> str:
    return re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", " ", text)


def pdf_text(path: Path) -> str:
    reader = PdfReader(str(path))
    return clean_pdf_text("\n".join(page.extract_text() or "" for page in reader.pages))


def first_number_by_patterns(text: str, patterns: Iterable[str]) -> float | None:
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE | re.S)
        if match:
            return float(match.group(1))
    return None


def extract_pdf_fields(path: Path) -> dict[str, float | None]:
    text = pdf_text(path)
    return {
        field: first_number_by_patterns(text, patterns)
        for field, patterns in PDF_FIELD_PATTERNS.items()
    }


def read_instrument_table(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix in {".xls", ".csv", ".tsv", ".txt"}:
        # The tested Coulter ".xls" files are tab-delimited text exports.
        return pd.read_csv(path, sep="\t", engine="python")
    if suffix == ".xlsx":
        return pd.read_excel(path)
    raise ValueError(f"Unsupported particle data file: {path}")


def read_particle_diameters(path: Path) -> pd.Series:
    df = read_instrument_table(path)
    if df.empty:
        return pd.Series(dtype=float)
    return pd.to_numeric(df.iloc[:, 0], errors="coerce").dropna()


def safe_ratio(numerator: float, denominator: float) -> float:
    return numerator / denominator if denominator else math.nan


def xls_stats(path: Path) -> dict[str, float | int]:
    x = read_particle_diameters(path)
    if x.empty:
        return {
            "particle_count_xls": 0,
            "mean_diameter_nm_xls": math.nan,
            "median_diameter_d50_nm_xls": math.nan,
            "d10_nm_xls": math.nan,
            "d90_nm_xls": math.nan,
            "d90_d10_ratio_xls": math.nan,
            "span_xls": math.nan,
            "std_diameter_nm_xls": math.nan,
            "cv_diameter_xls": math.nan,
            "skewness_xls": math.nan,
            "kurtosis_excess_xls": math.nan,
            "min_diameter_nm_xls": math.nan,
            "max_diameter_nm_xls": math.nan,
        }

    mean = float(x.mean())
    median = float(x.quantile(0.50))
    d10 = float(x.quantile(0.10))
    d90 = float(x.quantile(0.90))
    std = float(x.std(ddof=1))
    centered = x - mean
    m2 = float((centered**2).mean())
    m3 = float((centered**3).mean())
    m4 = float((centered**4).mean())
    skew = m3 / (m2 ** 1.5) if m2 > 0 else math.nan
    kurtosis_excess = m4 / (m2**2) - 3 if m2 > 0 else math.nan

    return {
        "particle_count_xls": int(x.size),
        "mean_diameter_nm_xls": mean,
        "median_diameter_d50_nm_xls": median,
        "d10_nm_xls": d10,
        "d90_nm_xls": d90,
        "d90_d10_ratio_xls": safe_ratio(d90, d10),
        "span_xls": safe_ratio(d90 - d10, median),
        "std_diameter_nm_xls": std,
        "cv_diameter_xls": safe_ratio(std, mean),
        "skewness_xls": skew,
        "kurtosis_excess_xls": kurtosis_excess,
        "min_diameter_nm_xls": float(x.min()),
        "max_diameter_nm_xls": float(x.max()),
    }


def infer_sample_number(stem: str) -> int | None:
    dilution_match = re.search(r"(?:^|[-_ .])(\d+)[-_ .]*\d+\s*X(?:$|[-_ .])", stem, re.IGNORECASE)
    if dilution_match:
        return int(dilution_match.group(1))
    leading_match = re.match(r"^(\d+)(?:[-_ .]|$)", stem)
    if leading_match:
        return int(leading_match.group(1))
    return None


def collect_pairs(input_dir: Path, recursive: bool) -> list[tuple[int | None, Path, Path]]:
    groups: dict[tuple[str, str], dict[str, Path]] = {}
    files = input_dir.rglob("*") if recursive else input_dir.iterdir()
    for file in files:
        if not file.is_file():
            continue
        suffix = file.suffix.lower()
        if suffix not in {".pdf", *PARTICLE_DATA_EXTENSIONS}:
            continue
        parent_key = str(file.parent.relative_to(input_dir)) if recursive else "."
        ext_key = "pdf" if suffix == ".pdf" else "particle"
        key = (parent_key, file.stem)
        groups.setdefault(key, {})
        groups[key].setdefault(ext_key, file)

    pairs: list[tuple[int | None, Path, Path]] = []
    for (_parent_key, stem), item in sorted(
        groups.items(),
        key=lambda value: (value[0][0], infer_sample_number(value[0][1]) or 10**9, value[0][1]),
    ):
        if "pdf" in item and "particle" in item:
            pairs.append((infer_sample_number(stem), item["pdf"], item["particle"]))
    return pairs


def write_xlsx(rows: list[dict[str, object]], out: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Coulter summary"
    headers = list(rows[0].keys()) if rows else []
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx, header in enumerate(headers, start=1):
        width = min(max(len(str(header)) + 2, 12), 56)
        for cell in ws[get_column_letter(col_idx)]:
            if cell.value is not None:
                width = min(max(width, len(str(cell.value)) + 2), 56)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)


def build_rows(input_dir: Path, recursive: bool) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for sample_no, pdf, particle_file in collect_pairs(input_dir, recursive):
        sample_id = f"{pdf.parent.name}-{pdf.stem}"
        row: dict[str, object] = {
            "sample_id": sample_id,
            "folder_name": pdf.parent.name,
            "file_stem": pdf.stem,
            "sample_number": sample_no,
            "experiment_name": "",
            "sample_name": "",
            "pdf_path": str(pdf),
            "particle_data_path": str(particle_file),
        }
        row.update(extract_pdf_fields(pdf))
        row.update(xls_stats(particle_file))
        rows.append(row)
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Extract one-wide-table Coulter PDF/XLS summary features for CQA data mining."
    )
    parser.add_argument("--input-dir", required=True, help="Folder containing paired Coulter PDF and particle data files.")
    parser.add_argument("--output-dir", required=True, help="Folder for output CSV/XLSX summary files.")
    parser.add_argument("--recursive", action="store_true", help="Search input folder recursively.")
    parser.add_argument("--prefix", default="coulter_summary_pilot", help="Output filename prefix.")
    args = parser.parse_args()

    input_dir = Path(args.input_dir)
    output_dir = Path(args.output_dir)
    rows = build_rows(input_dir, recursive=args.recursive)
    if not rows:
        raise SystemExit("No paired Coulter PDF and particle data files found.")

    output_dir.mkdir(parents=True, exist_ok=True)
    csv_out = output_dir / f"{args.prefix}.csv"
    xlsx_out = output_dir / f"{args.prefix}.xlsx"
    pd.DataFrame(rows).to_csv(csv_out, index=False, encoding="utf-8-sig")
    write_xlsx(rows, xlsx_out)

    print(csv_out)
    print(xlsx_out)
    print(f"rows={len(rows)}")


if __name__ == "__main__":
    main()
